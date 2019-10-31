import json
import urllib

import openpyxl
import time
import os
import pandas as pd
from datetime import datetime
from urllib import request
from urllib.parse import quote

class getpoi:
    #结果输出路径
    output_path = "E:\\PycharmProjects\\GaoDeMap\\"
    #POI分类编码表位置
    path_class = "E:\\PycharmProjects\\GaoDeMap\\amap_poicode.xlsx"
    #高德地图API_kEY
    amap_web_key = 'c9729799c0be9ea72e97081a652366a5'
    poi_search_url = "https://restapi.amap.com/v3/place/text?key=%s&extensions=all&keywords=&types=%s&city=%s&citylimit=true&offset=25&page=%s&output=json"
    #要搜索城市的名称
    cityname = '沈阳'
    #要搜索POI的区县名
    areas = ['和平区', '沈河区', '皇姑区', '大东区', '铁西区', '浑南区', '于洪区', '沈北新区', '苏家屯区', '辽中区', '新民市', '康平县', '法库县']
    totalcontent = {}

    def __init__(self):
        data_class = self.getclass()
        for type_class in data_class:
            for area in self.areas:
                page = 1;
                if type_class['type_num'] / 10000 < 10:
                    classtype = str('0') + str(type_class['type_num'])
                else:
                    classtype = str(type_class['type_num'])
                while True:
                    if classtype[-4:] == "0000":
                        break;
                    poidata = self.get_poi(classtype, area, page);
                    poidata = json.loads(poidata)

                    if poidata['count'] == "0":
                        break;
                    else:
                        poilist = self.hand(poidata)
                        print("area：" + area + "  type：" + classtype + "  page：第" + str(page) + "页  count：" + poidata[
                            'count'] + "poilist:")
                        page += 1
                        for pois in poilist:
                            if classtype[0:2] in self.totalcontent.keys():
                                pois['bigclass'] = type_class['bigclass']
                                pois['midclass'] = type_class['midclass']
                                pois['smallclass'] = type_class['smallclass']
                                list_total = self.totalcontent[classtype[0:2]]
                                list_total.append(pois)
                            else:
                                self.totalcontent[classtype[0:2]] = []
                                pois['bigclass'] = type_class['bigclass']
                                pois['midclass'] = type_class['midclass']
                                pois['smallclass'] = type_class['smallclass']
                                self.totalcontent[classtype[0:2]].append(pois)
        for content in self.totalcontent:
            self.writeexcel(self.totalcontent[content], content)

    def writeexcel(self, data, classname):
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = classname
        # 第一行(列标题)
        sheet.cell(row=1, column=1).value = 'x'
        sheet.cell(row=1, column=2).value = 'y'
        sheet.cell(row=1, column=3).value = 'name'
        sheet.cell(row=1, column=4).value = 'address'
        sheet.cell(row=1, column=5).value = 'adname'
        sheet.cell(row=1, column=6).value = 'smallclass'
        sheet.cell(row=1, column=7).value = 'typecode'
        sheet.cell(row=1, column=8).value = 'midclass'
        classname = data[0]['bigclass']
        for i in range(1, len(data)):
            sheet.cell(row=i + 1, column=1, value=data[i]['lng'])
            sheet.cell(row=i + 1, column=2, value=data[i]['lat'])
            sheet.cell(row=i + 1, column=3, value=data[i]['name'])
            sheet.cell(row=i + 1, column=4, value=str(data[i]['address']))
            sheet.cell(row=i + 1, column=5, value=str(data[i]['adname']))
            sheet.cell(row=i + 1, column=6, value=data[i]['smallclass'])
            sheet.cell(row=i + 1, column=7, value=data[i]['classname'])
            sheet.cell(row=i + 1, column=8, value=data[i]['midclass'])
        book.save(self.output_path + self.cityname + '_' + classname + '.xlsx')
        print("Write Complete")

    def hand(self, poidate):
        pois = poidate['pois']
        poilist = []
        for i in range(len(pois)):
            content = {}
            content['lng'] = float(str(pois[i]['location']).split(",")[0])
            content['lat'] = float(str(pois[i]['location']).split(",")[1])
            content['name'] = pois[i]['name']
            content['address'] = pois[i]['address']
            content['tel'] = pois[i]['tel']
            content['adname'] = pois[i]['adname']
            content['classname'] = pois[i]['typecode']
            poilist.append(content)
        return poilist

    def readfile(self, readfilename, sheetname):
        data = pd.read_excel(readfilename, sheet_name=sheetname)
        return data

    def getclass(self):
        readcontent = self.readfile(self.path_class, "POI分类与编码（中英文）")
        data = []
        for num in range(readcontent.shape[0]):
            content = {}
            content['type_num'] = readcontent.iloc[num]['NEW_TYPE']
            content['bigclass'] = readcontent.iloc[num]['大类']
            content['midclass'] = readcontent.iloc[num]['中类']
            content['smallclass'] = readcontent.iloc[num]['小类']
            data.append(content)
        return data

    def get_poi(self, keywords, city, page):
        poiurl = self.poi_search_url % (self.amap_web_key, keywords, quote(city), page)
        data = ''
        with urllib.request.urlopen(poiurl) as f:
            data = f.read().decode('utf8')
        return data


if __name__ == "__main__":
    gp = getpoi()






